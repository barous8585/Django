import csv
import io
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone
from datetime import timedelta
from django.db.models import Count, Sum
from core.models import Conteneur, Participation, Transaction, Portefeuille
import json
from decimal import Decimal

# Convertisseur JSON pour les Decimal
class DecimalEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, Decimal):
            return float(obj)
        return super(DecimalEncoder, self).default(obj)

def dashboard_view(request):
    # Export functionality
    export_type = request.GET.get('export')
    if export_type:
        return handle_export(export_type)
    
    # Statistics
    conteneurs = Conteneur.objects.filter(annule=False)
    participations = Participation.objects.all()
    transactions = Transaction.objects.all()
    
    total_collecte = conteneurs.aggregate(total=Sum('montant_actuel'))['total']
    stats = {
        'total_conteneurs': conteneurs.count(),
        'total_participants': participations.values('utilisateur').distinct().count(),
        'total_collecte': float(total_collecte) if total_collecte else 0,
        'en_attente': participations.filter(valide=False).count(),
        'validees': participations.filter(valide=True).count(),
    }
    
    # Chart data: Progression des conteneurs
    conteneur_labels = []
    conteneur_data = []
    collecte_data = []
    
    for c in conteneurs[:6]:
        conteneur_labels.append(c.nom[:20])
        conteneur_data.append(round(c.get_progression(), 2))
        collecte_data.append(float(c.montant_actuel))
    
    # Evolution des participations (7 derniers jours)
    today = timezone.now()
    evolution_labels = []
    evolution_data = []
    
    for i in range(6, -1, -1):
        date = today - timedelta(days=i)
        label = date.strftime('%d/%m')
        count = participations.filter(
            date_participation__date=date.date()
        ).count()
        evolution_labels.append(label)
        evolution_data.append(count)
    
    # Transactions récentes
    transactions_recentes = transactions.order_by('-date_transaction')[:6]
    
    context = {
        'title': '📊 Dashboard Admin',
        'description': 'Vue d\'ensemble de la plateforme Tontine Digitale',
        'stats': stats,
        'conteneur_labels': json.dumps(conteneur_labels, cls=DecimalEncoder),
        'conteneur_data': json.dumps(conteneur_data, cls=DecimalEncoder),
        'collecte_data': json.dumps(collecte_data, cls=DecimalEncoder),
        'evolution_labels': json.dumps(evolution_labels, cls=DecimalEncoder),
        'evolution_data': json.dumps(evolution_data, cls=DecimalEncoder),
        'transactions_recentes': transactions_recentes,
    }
    
    return render(request, 'dashboard.html', context)


def handle_export(export_type):
    if export_type == 'conteneurs':
        return export_conteneurs_csv()
    elif export_type == 'participations':
        return export_participations_csv()
    elif export_type == 'transactions':
        return export_transactions_csv()
    elif export_type == 'all':
        return export_all_excel()
    return HttpResponse("Type d'export non reconnu", status=400)


def export_conteneurs_csv():
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="conteneurs.csv"'
    response.write('\ufeff')
    
    writer = csv.writer(response)
    writer.writerow(['ID', 'Nom', 'Objectif', 'Devise', 'Montant Actuel', 'Progression (%)', 'Étape', 'Date Création'])
    
    for c in Conteneur.objects.filter(annule=False):
        writer.writerow([
            c.id,
            c.nom,
            c.objectif,
            c.devise,
            c.montant_actuel,
            c.get_progression(),
            c.get_etape_display(),
            c.date_creation.strftime('%Y-%m-%d %H:%M')
        ])
    
    return response


def export_participations_csv():
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="participations.csv"'
    response.write('\ufeff')
    
    writer = csv.writer(response)
    writer.writerow(['ID', 'Conteneur', 'Utilisateur', 'Téléphone', 'Montant', 'Référence', 'Validé', 'Date'])
    
    for p in Participation.objects.all():
        writer.writerow([
            p.id,
            p.conteneur.nom,
            p.utilisateur.username,
            p.utilisateur.telephone,
            p.montant,
            p.reference_paiement,
            'Oui' if p.valide else 'Non',
            p.date_participation.strftime('%Y-%m-%d %H:%M')
        ])
    
    return response


def export_transactions_csv():
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="transactions.csv"'
    response.write('\ufeff')
    
    writer = csv.writer(response)
    writer.writerow(['ID', 'Type', 'Utilisateur', 'Montant', 'Conteneur', 'Description', 'Date'])
    
    for t in Transaction.objects.all():
        writer.writerow([
            t.id,
            t.get_type_transaction_display(),
            t.portefeuille.utilisateur.telephone,
            t.montant,
            t.conteneur.nom if t.conteneur else '-',
            t.description,
            t.date_transaction.strftime('%Y-%m-%d %H:%M')
        ])
    
    return response


def export_all_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        
        wb = openpyxl.Workbook()
        
        # Sheet 1: Conteneurs
        ws1 = wb.active
        ws1.title = "Conteneurs"
        headers = ['ID', 'Nom', 'Objectif', 'Devise', 'Montant Actuel', 'Progression (%)', 'Étape', 'Date Création']
        ws1.append(headers)
        
        for cell in ws1[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        
        for c in Conteneur.objects.filter(annule=False):
            ws1.append([
                c.id, c.nom, float(c.objectif), c.devise, float(c.montant_actuel),
                c.get_progression(), c.get_etape_display(), c.date_creation
            ])
        
        # Sheet 2: Participations
        ws2 = wb.create_sheet("Participations")
        headers = ['ID', 'Conteneur', 'Utilisateur', 'Téléphone', 'Montant', 'Référence', 'Validé', 'Date']
        ws2.append(headers)
        
        for cell in ws2[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        
        for p in Participation.objects.all():
            ws2.append([
                p.id, p.conteneur.nom, p.utilisateur.username, p.utilisateur.telephone,
                float(p.montant), p.reference_paiement, 'Oui' if p.valide else 'Non', p.date_participation
            ])
        
        # Sheet 3: Transactions
        ws3 = wb.create_sheet("Transactions")
        headers = ['ID', 'Type', 'Utilisateur', 'Montant', 'Description', 'Date']
        ws3.append(headers)
        
        for cell in ws3[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        
        for t in Transaction.objects.all():
            ws3.append([
                t.id, t.get_type_transaction_display(), t.portefeuille.utilisateur.telephone,
                float(t.montant), t.description, t.date_transaction
            ])
        
        # Save to response
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="tontine_digitale_export.xlsx"'
        return response
        
    except ImportError:
        return HttpResponse(
            "openpyxl n'est pas installé. Utilisez: pip install openpyxl",
            status=500
        )
